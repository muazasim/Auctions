from django.contrib.auth import authenticate, login, logout
from django.db import IntegrityError
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
import openpyxl
import os
import pandas as pd
from .models import User, ActiveListing, Comment, Bid, Watchlist


def index(request):
    empty = False
    Listing = ActiveListing.objects.all()
    comments = Comment.objects.all()
    Bids = Bid.objects.all()
    if not Listing:
        empty = True
    return render(request, "auctions/index.html", {
        'Listing': Listing,
        'comments': comments,
        'Bids': Bids,
        'zero': 0,
        'head': 'Active Listings',
        'maxbidowner': '',
        'empty': empty

    })


def login_view(request):
    if request.method == "POST":

        # Attempt to sign user in
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        # Check if authentication successful
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse("index"))
        else:
            return render(request, "auctions/login.html", {
                "message": "Invalid username and/or password."
            })
    else:
        return render(request, "auctions/login.html")


def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse("index"))


def register(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]

        # Ensure password matches confirmation
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        if password != confirmation:
            return render(request, "auctions/register.html", {
                "message": "Passwords must match."
            })

        # Attempt to create new user
        try:
            user = User.objects.create_user(username, email, password)
            user.save()
        except IntegrityError:
            return render(request, "auctions/register.html", {
                "message": "Username already taken."
            })
        login(request, user)
        return HttpResponseRedirect(reverse("index"))
    else:
        return render(request, "auctions/register.html")


def addlisting(request):
    categories = ['SUV', 'OFROAD', 'LUXURY',
                  'SEDAN', 'HATCH-BACK', 'SPORTS', 'MUSCLE']
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    else:

        if request.method == "POST":
            title1 = request.POST['title']
            category1 = request.POST['category']
            startbid1 = request.POST['startbid']
            url1 = request.POST['url']
            currentbid1 = 0
            owner = ""
            discription1 = request.POST['discription']
            # listing = ActiveListing.objects.add(
            #     title, discription, url, category, startbid, request.user)
            ActiveListing.objects.create(
                title=title1, discription=discription1, url=url1, startbid=startbid1, category=category1, Listingowner=request.user, maxbid=currentbid1, maxbidonwer=request.user)
        return render(request, "auctions/addlisting.html", {
            "categories": categories
        })


def addcomment(request):
    if request.method == "POST":
        Comment.objects.create(
            comment=request.POST['comment'],
            commentlist=request.POST['listid1'],
            commentowner=request.user
        )
    return HttpResponseRedirect(reverse("index"))


def addbid(request):
    if request.method == "POST":
        list = ActiveListing.objects.get(id=request.POST['listid2'])
        bid = int(request.POST['bid'])
        user = request.user
        if list.startbid <= bid:
            if list.maxbid <= bid:
                list.maxbid = request.POST['bid']
                print(list.maxbid)
                list.maxbidonwer = request.user
                list.save()
                print(list.maxbid)
            Bid.objects.create(
                currentbid=request.POST['bid'],
                biddedlist=request.POST['listid2'],
                bidowner=request.user
            )

    return HttpResponseRedirect(reverse("index"))


def watchlist(request):
    userwatchlist = []
    listing = []
    empty = False
    userwatchlist = Watchlist.objects.filter(user=request.user)
    Activelists = ActiveListing.objects.all()

    for i in userwatchlist:
        for j in Activelists:
            if j.id == i.list:
                listing.append(j)
                print(j)
    if request.method == 'POST':
        listid = request.POST['listid3']
        user1 = request.user
        res = {}
        res = Watchlist.objects.filter(user=user1, list=listid)

        if not res:
            Watchlist.objects.create(
                user=user1,
                list=listid
            )
        else:
            res = Watchlist.objects.filter(user=user1, list=listid).delete()
        return HttpResponseRedirect(reverse("index"))

    else:
        comments = Comment.objects.all()
        Bids = Bid.objects.all()
        if not listing:
            empty = True
        return render(request, "auctions/watchlist.html", {
            "Listings": listing,
            'comments': comments,
            'empty': empty,
            'zero': 0,
            'Bids': Bids,

        })


def visitlisting(request, listid):
    comments = Comment.objects.all()
    Bids = Bid.objects.all()
    list1 = ActiveListing.objects.get(id=listid)
    user1 = request.user
    find = False
    endbid = True
    bid = False
    watchlist = {}
    watchlist = Watchlist.objects.filter(user=user1, list=listid)
    if list1.maxbid > 0:
        bid = True
    if not watchlist:
        find = True

    onwer = {}
    owner = ActiveListing.objects.filter(id=listid, Listingowner=user1)
    if not owner:
        endbid = False

    return render(request, "auctions/visitlisting.html", {
        "list": list1,
        'comments': comments,
        'Bids': Bids,
        'find': find,
        'endbid': endbid,
        'bid': bid


    })


def endbid(request, listid):
    if request.method == 'POST':
        list = ActiveListing.objects.get(id=listid)
        list.active = False
        list.save()
        return HttpResponseRedirect(reverse("index"))


def wonlisting(request):
    found = True
    Listings = ActiveListing.objects.filter(
        maxbidonwer=request.user, active=False)
    comments = Comment.objects.all()
    Bids = Bid.objects.all()
    if not Listings:
        found = False
    return render(request, "auctions/wonlisting.html", {
        'Listing': Listings,
        'comments': comments,
        'Bids': Bids,
        'zero': 0,
        'Found': found,
        'maxbidowner': ''

    })


def categories(request):
    categories = []
    empty = False
    listings = ActiveListing.objects.all()
    for list in listings:

        c = list.category
        if c not in categories:
            categories.append(c)
    if not categories:
        empty = True
    return render(request, "auctions/categories.html", {
        "categories": categories,
        'empty': empty

    })


def displaycategories(request, category):
    category1 = category
    Listing = ActiveListing.objects.filter(category=category1)
    comments = Comment.objects.all()
    Bids = Bid.objects.all()
    return render(request, "auctions/index.html", {
        'Listing': Listing,
        'comments': comments,
        'Bids': Bids,
        'zero': 0,
        'head': category1,
        'maxbidowner': ''

    })


def work(request):
    values = []  # to display table
    if request.method == 'POST':
        uploaded_file = request.FILES['myfile']
        wb = openpyxl.load_workbook(
            "media/company.xlsx", keep_vba=True, data_only=True)
        ws = wb.active
        workbook = wb.active
        # Load the Excel sheet
        # workbook = openpyxl.load_workbook('media/company.xlsx')

        # Select the worksheet
        # worksheet = workbook[worksheet_name]

        # Create a list to store the cell values

        # Loop through all the rows in the worksheet
        for row in workbook.iter_rows():
            # Create a list to store the row values
            row_values = []

        # Loop through all the cells in the row
        for cell in row:
            # Get the value of the cell
            cell_value = cell.value

            # Append the cell value to the row values list
            row_values.append(cell_value)

        # Append the row values list to the values list
        values.append(row_values)
        table_data = []
        for row in ws.iter_rows(values_only=True):
            table_data.append(row)

        df = pd.read_excel(uploaded_file, index_col=False)
        # df.drop(df.columns[0], axis=1, inplace=True)
        # print(df)
        html_table = df.to_html()
        amountcolumn = request.POST.getlist('amounts')
        labelcolumn = int(request.POST['label'])
        # print(labelcolumn)
        result = {}
        row = len(df)
        col = df.shape[1]
        # getting all labeles in a list
        labels = df[df.columns[labelcolumn-1]].drop_duplicates().copy()

        # REMOVING LABELS FROM THE DATAFRAME
        df = df.drop(columns=df.columns[0])
        # print(df)  # after removing label column

        # df1 = df.drop(df.columns[labelcolumn-1], axis=1)
        df = df.apply(pd.to_numeric, errors='coerce')

        # GETTING SUM OF EACH ROW
        sum = df.sum(axis=1, skipna=True)

        # print(sum)
        for i in sum:
            print(i)
        # print(len(labels))

        # TO GET THE INDEX OF VERY FIRST THAT CONTAINS A NUMERIC VALUE
        for index, row in df.iterrows():
            # use pd.to_numeric() to convert the row to numeric values, ignoring non-numeric values with errors='coerce'
            numeric_row = pd.to_numeric(row, errors='coerce')
        # check if any values in the numeric_row are not NaN
            if not pd.isna(numeric_row).all():
                # if there is at least one non-NaN value, then this is the first row with a numeric value
                # first_numeric_row = df.iloc[index]
                # print the first numeric row and exit the loop
                # print("The first row with a numeric value is:", first_numeric_row)
                # break

                first_numeric_row = df.iloc[index]
                # print the index of the first numeric row and exit the loop
                print("The index of the first row with a numeric value is:", index)
                break

        # return render(request, 'auctions/sheet.html', {'html_table': html_table})

    return render(request, 'auctions/displaysheet.html', {'table_data': values})
